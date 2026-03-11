
import csv
import os
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

try:
    import pandas as pd  # type: ignore
except Exception:  # pragma: no cover
    pd = None

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None

COLUMN_CODE = "Q3"
COLUMN_NAME = "Q4"
DATA_START_OFFSET = 2
MODULE_LEADERS_CANDIDATES = ("Module leader(s)", "Module leader", "module leader(s)", "module leader")
LOCATION_CANDIDATES = ("Location", "location")

SKILL_COL_START = "Q5_1"
SKILL_COL_END = "Q23"
CAREER_COL_START = "Q25_1"
CAREER_COL_END = "Q25_6"


def read_rows(file_path: str):
    """
    Returns a pandas.DataFrame when pandas can read the file; otherwise a list[dict] for CSV.
    Mirrors the behavior in the original Tkinter app.
    """
    ext = os.path.splitext(file_path)[1].lower()

    if pd is not None:
        try:
            if ext in (".xlsx", ".xls"):
                df = pd.read_excel(file_path)
            else:
                df = pd.read_csv(file_path)
            return df
        except Exception:
            pass

    if ext == ".xlsx" and load_workbook is not None:
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active
            rows = list(ws.values)
            if not rows:
                return []
            headers = ["" if h is None else str(h).strip() for h in rows[0]]
            data = []
            for r in rows[1:]:
                row = {}
                for i, key in enumerate(headers):
                    if not key:
                        continue
                    row[key] = r[i] if i < len(r) else None
                data.append(row)
            return data
        except Exception:
            pass

    if ext in (".xlsx", ".xls"):
        if ext == ".xls":
            raise RuntimeError("No Excel reader available for .xls files. Please install pandas.")
        raise RuntimeError("No Excel reader available. Please install openpyxl.")

    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            with open(file_path, "r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            return rows
        except Exception:
            continue
    raise RuntimeError("Unable to read CSV file. Please check encoding.")


def parse_score(value: Any) -> Optional[int]:
    if value is None:
        return None

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and value != value:  # NaN
            return None
        return int(round(value))

    raw = str(value).strip()
    if not raw:
        return None

    if raw.isdigit():
        return int(raw)

    try:
        num = float(raw)
        if num != num:  # NaN
            return None
        return int(round(num))
    except ValueError:
        pass

    normalized = raw.lower()
    if normalized == "no":
        return 0
    if normalized == "somewhat":
        return 1
    if normalized == "yes":
        return 2
    return None


def get_skill_columns(columns: List[str]) -> List[str]:
    if SKILL_COL_START not in columns:
        return []
    start_idx = columns.index(SKILL_COL_START)
    if SKILL_COL_END in columns:
        end_idx = columns.index(SKILL_COL_END)
        return columns[start_idx : end_idx + 1]
    return columns[start_idx:]


def get_career_columns(columns: List[str]) -> List[str]:
    if CAREER_COL_START not in columns:
        return []
    start_idx = columns.index(CAREER_COL_START)
    if CAREER_COL_END in columns:
        end_idx = columns.index(CAREER_COL_END)
        return columns[start_idx : end_idx + 1]
    return columns[start_idx:]


_SKILL_CATEGORIES = [
    "Foundational Skills",
    "Communication",
    "Leadership",
    "Adaptability",
    "Creativity",
    "Digital Competency",
]


def split_skill_category(label: str) -> Tuple[str, str]:
    raw = label.strip()
    for category in _SKILL_CATEGORIES:
        if raw.lower().startswith(category.lower()):
            skill = raw[len(category):].lstrip(": -").strip() or raw
            return category, skill
    # fallback: try "Category: Skill"
    if ":" in raw:
        cat, skill = raw.split(":", 1)
        return cat.strip(), skill.strip()
    return "Other", raw


def build_skills(columns: List[str], label_row: Any, row: Any) -> List[Dict[str, Any]]:
    if not columns:
        return []
    items: List[Dict[str, Any]] = []
    for col in columns:
        value = row.get(col) if hasattr(row, "get") else row[col]
        label = ""
        if label_row is not None:
            label_val = label_row.get(col) if hasattr(label_row, "get") else label_row[col]
            if label_val is not None:
                label = str(label_val).strip()

        score = parse_score(value)
        if score is None:
            continue

        label = label or col
        category, skill = split_skill_category(label)
        display_label = skill if category in _SKILL_CATEGORIES else label.split(":", 1)[0].strip()
        items.append(
            {"label": display_label, "full_label": label, "category": category, "skill": skill, "score": score}
        )
    return items


def build_career_pathways(columns: List[str], label_row: Any, row: Any) -> List[Dict[str, Any]]:
    if not columns:
        return []
    items: List[Dict[str, Any]] = []
    for col in columns:
        value = row.get(col) if hasattr(row, "get") else row[col]
        score = parse_score(value)
        if score is None:
            continue

        label = col
        if label_row is not None:
            label_val = label_row.get(col) if hasattr(label_row, "get") else label_row[col]
            if label_val is not None:
                label = str(label_val).strip()

        career = label.split(" - ", 1)[1].strip() if " - " in label else label.strip()
        items.append({"career": career, "score": score})
    return items


def _pick_optional_text(row: Any, candidates: Tuple[str, ...]) -> str:
    """Read an optional text value from a row using flexible column name matching."""
    if row is None:
        return ""

    # Exact match first.
    for key in candidates:
        val = row.get(key) if hasattr(row, "get") else None
        if val is not None and str(val).strip():
            return str(val).strip()

    # Case/spacing-insensitive fallback for template variations.
    if hasattr(row, "keys"):
        normalized_map = {
            str(k).strip().lower().replace(" ", "").replace("_", ""): k
            for k in row.keys()
            if k is not None
        }
        for key in candidates:
            norm = str(key).strip().lower().replace(" ", "").replace("_", "")
            actual = normalized_map.get(norm)
            if actual is None:
                continue
            val = row.get(actual) if hasattr(row, "get") else None
            if val is not None and str(val).strip():
                return str(val).strip()
    return ""


def build_module(
    code: str,
    name: str,
    skills=None,
    careers=None,
    module_leaders: str = "",
    location: str = "",
) -> Dict[str, Any]:
    return {
        "code": code,
        "name": name,
        "skills": skills or [],
        "careers": careers or [],
        "information": "",
        "assessment": "",
        "assessment_type": "",
        "module_leaders": module_leaders,
        "location": location,
    }


def load_template_modules(file_path: str) -> List[Dict[str, Any]]:
    data = read_rows(file_path)

    modules: List[Dict[str, Any]] = []
    label_row = None

    if pd is not None and hasattr(data, "columns"):
        if COLUMN_CODE not in data.columns or COLUMN_NAME not in data.columns:
            raise ValueError(f"Missing columns: {COLUMN_CODE} or {COLUMN_NAME}")

        skill_columns = get_skill_columns(list(data.columns))
        career_columns = get_career_columns(list(data.columns))
        label_row = data.iloc[0] if len(data) >= 1 else None

        for idx, row in data.iterrows():
            if idx < DATA_START_OFFSET:
                continue
            code = "" if row.get(COLUMN_CODE) is None else str(row.get(COLUMN_CODE)).strip()
            name = "" if row.get(COLUMN_NAME) is None else str(row.get(COLUMN_NAME)).strip()
            module_leaders = _pick_optional_text(row, MODULE_LEADERS_CANDIDATES)
            location = _pick_optional_text(row, LOCATION_CANDIDATES)
            skills = build_skills(skill_columns, label_row, row)
            careers = build_career_pathways(career_columns, label_row, row)
            modules.append(build_module(code, name, skills, careers, module_leaders=module_leaders, location=location))
        return modules

    # CSV fallback (list[dict])
    if not data:
        return []
    if COLUMN_CODE not in data[0] or COLUMN_NAME not in data[0]:
        raise ValueError(f"Missing columns: {COLUMN_CODE} or {COLUMN_NAME}")

    skill_columns = get_skill_columns(list(data[0].keys()))
    career_columns = get_career_columns(list(data[0].keys()))
    label_row = data[0] if len(data) >= 1 else None

    for idx, row in enumerate(data):
        if idx < DATA_START_OFFSET:
            continue
        code = (row.get(COLUMN_CODE) or "").strip()
        name = (row.get(COLUMN_NAME) or "").strip()
        module_leaders = _pick_optional_text(row, MODULE_LEADERS_CANDIDATES)
        location = _pick_optional_text(row, LOCATION_CANDIDATES)
        skills = build_skills(skill_columns, label_row, row)
        careers = build_career_pathways(career_columns, label_row, row)
        modules.append(build_module(code, name, skills, careers, module_leaders=module_leaders, location=location))
    return modules


def apply_adjustments(modules: List[Dict[str, Any]], adjustment: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    adjustment schema (minimal):
    {
      "modules": {
         "<code>": {"name": "...", "information": "...", "assessment": "...", "assessment_type": "..."}
      }
    }
    """
    by_code = (adjustment or {}).get("modules", {}) if isinstance(adjustment, dict) else {}
    out: List[Dict[str, Any]] = []
    for m in modules:
        code = (m.get("code") or "").strip()
        patch = by_code.get(code, {}) if isinstance(by_code, dict) else {}
        merged = dict(m)
        if isinstance(patch, dict):
            for k in ("code", "name", "information", "assessment", "assessment_type", "module_leaders", "location"):
                if k in patch and patch[k] is not None:
                    merged[k] = str(patch[k])
        out.append(merged)
    return out

